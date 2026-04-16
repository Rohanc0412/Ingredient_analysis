import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from helpers.excel_writer import apply_output_sheet_layout, find_data_sheet, load_workbook_context


class ExcelWriterTests(unittest.TestCase):
    def test_accepts_generic_single_sheet_template(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "template.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws.append(
                [
                    "SL No",
                    "Year",
                    "Authors",
                    "Title",
                    "Journal / Source",
                    "Primary Ingredient",
                ]
            )
            wb.save(path)

            ctx = load_workbook_context(path)

            # data_sheet should be set and headers read as-is from the template
            self.assertIsNotNone(ctx.data_sheet)
            self.assertEqual("SL No", ctx.headers[0])
            self.assertIn("Title", ctx.headers)
            self.assertIn("Primary Ingredient", ctx.headers)

    def test_find_data_sheet_returns_first_non_index_sheet(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "My Data Sheet"
        ws.append(["Col A", "Col B", "Col C"])

        data_ws = find_data_sheet(wb)

        self.assertEqual("My Data Sheet", data_ws.title)

    def test_find_data_sheet_skips_file_index(self):
        from helpers.excel_writer import FILE_INDEX_SHEET
        wb = Workbook()
        # Rename active sheet to File Index
        ws_index = wb.active
        ws_index.title = FILE_INDEX_SHEET
        ws_index.append(["Ref #", "Path", "SHA256"])
        # Add a real data sheet
        ws_data = wb.create_sheet("Extraction Results")
        ws_data.append(["Ingredient", "Year", "Study Type"])

        data_ws = find_data_sheet(wb)

        self.assertEqual("Extraction Results", data_ws.title)

    def test_apply_output_sheet_layout_marks_string_cells_as_text(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Ingredient"
        ws["A2"] = "- Curcumin"
        ws["B2"] = 42

        apply_output_sheet_layout(ws, min_row=1)

        self.assertEqual("@", ws["A2"].number_format)
        self.assertNotEqual("@", ws["B2"].number_format)


if __name__ == "__main__":
    unittest.main()
