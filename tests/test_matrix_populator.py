import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from populator_ingredient_matrix.cli import extract_answer, load_records, normalize_key, populate_workbook


class MatrixPopulatorTests(unittest.TestCase):
    def test_load_records_extracts_supported_field_shapes(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            payload = {
                "ingredient": "Example Ingredient",
                "fields": {
                    "Ingredient Category (Botanical, Vitamin, Mineral, Probiotic, Peptide, Fiber, etc.)": {
                        "category": "- Botanical",
                        "sources": [],
                    },
                    "Evidence of GLP-1 Interaction (Yes/No/Indirect)": {
                        "category": "- Indirect",
                        "sources": [],
                    },
                    "Clinical Evidence Level (Human RCT, Human Observational, Animal, In-vitro)": {
                        "category": "- Human RCT",
                        "sources": [],
                    },
                    "Number of Clinical Studies": {
                        "summary": "- More than 12 studies reported",
                        "study_count": 12,
                        "sources": [],
                    },
                    "Key Clinical Outcomes (Weight Loss %, Appetite Reduction, HbA1c, etc.)": {
                        "summary": "- Weight loss improved",
                        "study_count": 4,
                        "sources": [],
                    },
                    "Formulation Format (Capsule, Powder, Drink, Gummy, Sachet)": {
                        "category": "- Capsule, powder",
                        "sources": [],
                    },
                    "Average Effective Dose": {
                        "value": 1500,
                        "unit": "mg/day",
                        "context": "- Common regimen is split across meals",
                        "sources": [],
                    },
                    "Consumer Perceived Naturalness (High/Medium/Low)": {
                        "category": "- High",
                        "sources": [],
                    },
                    "Gender Skew": {
                        "category": "- Neutral",
                        "sources": [],
                    },
                    "Synergy Ingredients Commonly Used": {
                        "items": ["- Fiber", "- Chromium"],
                        "sources": [],
                    },
                },
            }
            (root / "example.json").write_text(json.dumps(payload), encoding="utf-8")

            records = load_records(root)

            self.assertEqual(1, len(records))
            record = records[0]
            self.assertEqual("Example Ingredient", record[normalize_key("Ingredient")])
            self.assertEqual(
                "- Botanical",
                record[normalize_key("Ingredient Category (Botanical, Vitamin, Mineral, Probiotic, Peptide, Fiber, etc.)")],
            )
            self.assertEqual(
                "- Indirect",
                record[normalize_key("Evidence of GLP-1 Interaction (Yes/No/Indirect)")],
            )
            self.assertEqual(
                "- Human RCT",
                record[normalize_key("Clinical Evidence Level (Human RCT, Human Observational, Animal, In-vitro)")],
            )
            self.assertEqual(
                "- More than 12 studies reported\n12",
                record[normalize_key("Number of Clinical Studies")],
            )
            self.assertEqual(
                "- Weight loss improved\n4",
                record[normalize_key("Key Clinical Outcomes (Weight Loss %, Appetite Reduction, HbA1c, etc.)")],
            )
            self.assertEqual(
                "- Capsule, powder",
                record[normalize_key("Formulation Format (Capsule, Powder, Drink, Gummy, Sachet)")],
            )
            self.assertEqual(
                "1500 mg/day\n- Common regimen is split across meals",
                record[normalize_key("Average Effective Dose")],
            )
            self.assertEqual(
                "- High",
                record[normalize_key("Consumer Perceived Naturalness (High/Medium/Low)")],
            )
            self.assertEqual("- Neutral", record[normalize_key("Gender Skew")])
            self.assertEqual(
                "- Fiber\n- Chromium",
                record[normalize_key("Synergy Ingredients Commonly Used")],
            )

    def test_extract_answer_recursively_reads_dynamic_shapes(self):
        value = {
            "result": {
                "headline": "Supports mitochondrial function",
                "details": {
                    "short_text": "Improves fatty acid transport",
                    "bullets": ["- Human data available", "- Favorable safety profile"],
                },
            },
            "sources": [{"id": 1}],
        }

        answer = extract_answer(value)

        self.assertIn("Supports mitochondrial function", answer)
        self.assertIn("Improves fatty acid transport", answer)
        self.assertIn("- Human data available", answer)
        self.assertIn("- Favorable safety profile", answer)

    def test_load_records_uses_template_header_for_ingredient_fallback(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            payload = {
                "ingredient": "Berberine",
                "fields": {
                    "Evidence Level": {"category": "Human"},
                },
            }
            (root / "example.json").write_text(json.dumps(payload), encoding="utf-8")

            records = load_records(root, template_headers=["Primary Ingredient", "Evidence Level"])

            self.assertEqual("Berberine", records[0][normalize_key("Primary Ingredient")])

    def test_populate_workbook_uses_detected_template_header_row(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            template_path = root / "template.xlsx"
            output_path = root / "output.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.title = "Matrix"
            ws["A1"] = "Weight Management Ingredient Matrix"
            ws.append([None, None])
            ws.append(["Primary Ingredient", "Evidence Level"])
            wb.save(template_path)

            populate_workbook(
                template_path,
                output_path,
                [{normalize_key("Primary Ingredient"): "Berberine", normalize_key("Evidence Level"): "Human RCT"}],
            )

            result_wb = load_workbook(output_path)
            result_ws = result_wb["Matrix"]

            self.assertEqual("Primary Ingredient", result_ws["A3"].value)
            self.assertEqual("Evidence Level", result_ws["B3"].value)
            self.assertEqual("Berberine", result_ws["A4"].value)
            self.assertEqual("Human RCT", result_ws["B4"].value)


if __name__ == "__main__":
    unittest.main()
