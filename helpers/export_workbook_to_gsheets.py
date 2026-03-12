from __future__ import annotations

import argparse
import math
import os
from json import JSONDecodeError
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Color
from openpyxl.utils import get_column_letter

try:
    from google.auth.transport.requests import Request
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from googleapiclient.discovery import build
    from googleapiclient.errors import HttpError
except ModuleNotFoundError:
    Credentials = None
    build = None
    Request = None
    InstalledAppFlow = None

    class HttpError(Exception):
        pass

try:
    from helpers.env import load_dotenv
    from helpers.logging_utils import get_logger
except ModuleNotFoundError:
    import sys

    ROOT = Path(__file__).resolve().parent.parent
    if str(ROOT) not in sys.path:
        sys.path.insert(0, str(ROOT))
    from helpers.env import load_dotenv
    from helpers.logging_utils import get_logger


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX_PATH = ROOT / "output" / "ingredient_analysis_output.xlsx"
GOOGLE_SHEETS_SCOPES = (
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
)
logger = get_logger(__name__, prefix="[ GSheets Export: ]")


def _safe_print(message: str) -> None:
    logger.info("%s", message)


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(prog="export_workbook_to_gsheets")
    parser.add_argument(
        "--xlsx-path",
        default="",
        help="Path to the source .xlsx workbook.",
    )
    parser.add_argument(
        "--spreadsheet-title",
        default="",
        help="Title for the created Google Sheets spreadsheet. Defaults to workbook stem with a timestamp.",
    )
    return parser


def _default_spreadsheet_title(xlsx_path: Path) -> str:
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    return f"{xlsx_path.stem} {ts}"


def _resolve_env_path(var_name: str, *, default: str = "") -> Path | None:
    raw = (os.environ.get(var_name) or default).strip()
    if not raw:
        return None
    path = Path(raw).expanduser()
    if not path.is_absolute():
        path = (ROOT / path).resolve()
    return path


def _load_credentials(client_secret_path: Path, token_path: Path):
    if Credentials is None or InstalledAppFlow is None or Request is None:
        raise RuntimeError(
            "Google API dependencies are not installed. Run: pip install -r requirements.txt"
        )

    creds = None
    try:
        if token_path.exists():
            creds = Credentials.from_authorized_user_file(str(token_path), GOOGLE_SHEETS_SCOPES)
    except (OSError, ValueError, JSONDecodeError) as exc:
        raise RuntimeError(
            f"Failed to load cached Google OAuth token: {token_path} ({type(exc).__name__}: {exc})"
        ) from exc

    try:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
            token_path.parent.mkdir(parents=True, exist_ok=True)
            token_path.write_text(creds.to_json(), encoding="utf-8")
            return creds
        if creds and creds.valid:
            return creds
    except Exception as exc:
        raise RuntimeError(
            f"Failed to refresh cached Google OAuth token: {token_path} ({type(exc).__name__}: {exc})"
        ) from exc

    try:
        flow = InstalledAppFlow.from_client_secrets_file(str(client_secret_path), GOOGLE_SHEETS_SCOPES)
        creds = flow.run_local_server(port=0)
        token_path.parent.mkdir(parents=True, exist_ok=True)
        token_path.write_text(creds.to_json(), encoding="utf-8")
        return creds
    except Exception as exc:
        raise RuntimeError(
            f"Failed to complete Google OAuth login using client secret: {client_secret_path} ({type(exc).__name__}: {exc})"
        ) from exc


def _build_services(credentials):
    if build is None:
        raise RuntimeError(
            "Google API dependencies are not installed. Run: pip install -r requirements.txt"
        )
    sheets_service = build("sheets", "v4", credentials=credentials)
    drive_service = build("drive", "v3", credentials=credentials)
    return sheets_service, drive_service


def _excel_width_to_pixels(width: float | None) -> int | None:
    if width is None:
        return None
    if width <= 0:
        return None
    # Approximate Excel's character-based width to Google Sheets pixel width.
    return max(21, int(math.ceil(float(width) * 7 + 5)))


def _argb_to_rgb(color_value: str) -> dict[str, float] | None:
    value = (color_value or "").strip()
    if len(value) == 8:
        value = value[2:]
    if len(value) != 6:
        return None
    try:
        r = int(value[0:2], 16) / 255.0
        g = int(value[2:4], 16) / 255.0
        b = int(value[4:6], 16) / 255.0
    except ValueError:
        return None
    return {"red": r, "green": g, "blue": b}


def _color_to_gsheets(color: Color | None) -> dict[str, float] | None:
    if color is None:
        return None
    rgb = getattr(color, "rgb", None)
    if rgb:
        return _argb_to_rgb(str(rgb))
    return None


def _horizontal_alignment(value: str | None) -> str | None:
    mapping = {
        "left": "LEFT",
        "center": "CENTER",
        "right": "RIGHT",
        "fill": "LEFT",
        "justify": "JUSTIFY",
        "centercontinuous": "CENTER",
        "distributed": "JUSTIFY",
        "general": None,
    }
    key = (value or "").strip().lower()
    return mapping.get(key)


def _vertical_alignment(value: str | None) -> str | None:
    mapping = {
        "top": "TOP",
        "center": "MIDDLE",
        "bottom": "BOTTOM",
        "justify": "MIDDLE",
        "distributed": "MIDDLE",
    }
    key = (value or "").strip().lower()
    return mapping.get(key)


def _border_style(value: str | None) -> str | None:
    mapping = {
        "thin": "SOLID",
        "medium": "SOLID_MEDIUM",
        "thick": "SOLID_THICK",
        "dashed": "DASHED",
        "dotted": "DOTTED",
        "double": "DOUBLE",
        "dashdot": "DASHED",
        "dashdotdot": "DASHED",
        "mediumdashed": "DASHED",
        "mediumdashdot": "DASHED",
        "mediumdashdotdot": "DASHED",
        "slantdashdot": "DASHED",
        "hair": "DOTTED",
    }
    key = (value or "").strip().lower()
    return mapping.get(key)


def _side_to_gsheets(side) -> dict[str, Any] | None:
    style = _border_style(getattr(side, "style", None))
    if not style:
        return None
    payload: dict[str, Any] = {"style": style}
    color = _color_to_gsheets(getattr(side, "color", None))
    if color:
        payload["color"] = color
    return payload


def _infer_number_format(format_code: str | None) -> tuple[str, str] | None:
    if not format_code:
        return None
    code = str(format_code).strip()
    if not code or code.lower() == "general":
        return None

    lowered = code.lower()
    if "%" in lowered:
        return ("PERCENT", code)
    if any(symbol in code for symbol in ("$", "₹", "€", "£", "¥")):
        return ("CURRENCY", code)
    if any(token in lowered for token in ("yy", "dd", "mm", "hh", "ss", "am/pm")):
        if any(token in lowered for token in ("hh", "ss", "am/pm")):
            return ("DATE_TIME", code)
        return ("DATE", code)
    if "0" in code or "#" in code:
        return ("NUMBER", code)
    return None


def _font_family(font) -> str | None:
    name = getattr(font, "name", None)
    return str(name) if name else None


def _cell_format(cell: Cell) -> dict[str, Any] | None:
    fmt: dict[str, Any] = {}

    font = cell.font
    if font:
        text_format: dict[str, Any] = {}
        family = _font_family(font)
        if family:
            text_format["fontFamily"] = family
        if getattr(font, "sz", None):
            text_format["fontSize"] = float(font.sz)
        if getattr(font, "b", False):
            text_format["bold"] = True
        if getattr(font, "i", False):
            text_format["italic"] = True
        if getattr(font, "u", None):
            text_format["underline"] = True
        foreground = _color_to_gsheets(getattr(font, "color", None))
        if foreground:
            text_format["foregroundColor"] = foreground
        if text_format:
            fmt["textFormat"] = text_format

    fill = cell.fill
    if fill and getattr(fill, "fill_type", None) == "solid":
        background = _color_to_gsheets(getattr(fill, "fgColor", None))
        if background:
            fmt["backgroundColor"] = background

    alignment = cell.alignment
    if alignment:
        horizontal = _horizontal_alignment(getattr(alignment, "horizontal", None))
        vertical = _vertical_alignment(getattr(alignment, "vertical", None))
        if horizontal:
            fmt["horizontalAlignment"] = horizontal
        if vertical:
            fmt["verticalAlignment"] = vertical
        fmt["wrapStrategy"] = "WRAP" if bool(getattr(alignment, "wrap_text", None) or getattr(alignment, "wrapText", None)) else "OVERFLOW_CELL"

    border = cell.border
    if border:
        borders: dict[str, Any] = {}
        for source_key, target_key in (
            ("left", "left"),
            ("right", "right"),
            ("top", "top"),
            ("bottom", "bottom"),
        ):
            side_payload = _side_to_gsheets(getattr(border, source_key, None))
            if side_payload:
                borders[target_key] = side_payload
        if borders:
            fmt["borders"] = borders

    number_format = _infer_number_format(cell.number_format)
    if number_format:
        fmt["numberFormat"] = {"type": number_format[0], "pattern": number_format[1]}

    return fmt or None


def _cell_value(cell: Cell) -> dict[str, Any]:
    value = cell.value
    if value is None:
        return {}
    if isinstance(value, bool):
        return {"boolValue": value}
    if isinstance(value, (int, float)):
        return {"numberValue": float(value)}
    if isinstance(value, Decimal):
        return {"numberValue": float(value)}
    if hasattr(value, "isoformat") and not isinstance(value, str):
        return {"stringValue": value.isoformat(sep=" ")}
    return {"stringValue": str(value)}


def _cell_data(cell: Cell) -> dict[str, Any]:
    data: dict[str, Any] = {}
    value = _cell_value(cell)
    if value:
        data["userEnteredValue"] = value
    cell_format = _cell_format(cell)
    if cell_format:
        data["userEnteredFormat"] = cell_format
    return data


def _sheet_update_request(ws, sheet_id: int) -> dict[str, Any]:
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        rows.append({"values": [_cell_data(cell) for cell in row]})

    return {
        "updateCells": {
            "start": {
                "sheetId": sheet_id,
                "rowIndex": 0,
                "columnIndex": 0,
            },
            "rows": rows,
            "fields": "userEnteredValue,userEnteredFormat",
        }
    }


def _dimension_requests(ws, sheet_id: int) -> list[dict[str, Any]]:
    requests: list[dict[str, Any]] = [
        {
            "updateSheetProperties": {
                "properties": {
                    "sheetId": sheet_id,
                    "gridProperties": {
                        "rowCount": max(1, ws.max_row),
                        "columnCount": max(1, ws.max_column),
                    },
                },
                "fields": "gridProperties.rowCount,gridProperties.columnCount",
            }
        }
    ]

    for column_index in range(1, ws.max_column + 1):
        dimension = ws.column_dimensions.get(get_column_letter(column_index))
        pixel_size = _excel_width_to_pixels(getattr(dimension, "width", None) if dimension else None)
        if pixel_size is None:
            continue
        requests.append(
            {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": sheet_id,
                        "dimension": "COLUMNS",
                        "startIndex": column_index - 1,
                        "endIndex": column_index,
                    },
                    "properties": {"pixelSize": pixel_size},
                    "fields": "pixelSize",
                }
            }
        )
    return requests


def _merge_requests(ws, sheet_id: int) -> list[dict[str, Any]]:
    requests: list[dict[str, Any]] = []
    for merged in ws.merged_cells.ranges:
        requests.append(
            {
                "mergeCells": {
                    "range": {
                        "sheetId": sheet_id,
                        "startRowIndex": merged.min_row - 1,
                        "endRowIndex": merged.max_row,
                        "startColumnIndex": merged.min_col - 1,
                        "endColumnIndex": merged.max_col,
                    },
                    "mergeType": "MERGE_ALL",
                }
            }
        )
    return requests


def _chunked(items: list[dict[str, Any]], size: int = 100) -> list[list[dict[str, Any]]]:
    return [items[idx: idx + size] for idx in range(0, len(items), size)]


def _get_spreadsheet_sheet_ids(sheets_service, spreadsheet_id: str) -> dict[str, int]:
    response = sheets_service.spreadsheets().get(
        spreadsheetId=spreadsheet_id,
        fields="sheets.properties(sheetId,title)",
    ).execute()
    return {
        sheet["properties"]["title"]: int(sheet["properties"]["sheetId"])
        for sheet in response.get("sheets", [])
    }


def _configure_sheets_structure(sheets_service, spreadsheet_id: str, workbook) -> dict[str, int]:
    metadata = sheets_service.spreadsheets().get(
        spreadsheetId=spreadsheet_id,
        fields="sheets.properties(sheetId,title,index,gridProperties)",
    ).execute()
    existing_sheets = metadata.get("sheets", [])
    requests: list[dict[str, Any]] = []

    if existing_sheets:
        first_sheet_id = int(existing_sheets[0]["properties"]["sheetId"])
        first_ws = workbook.worksheets[0]
        requests.append(
            {
                "updateSheetProperties": {
                    "properties": {
                        "sheetId": first_sheet_id,
                        "title": first_ws.title,
                        "index": 0,
                        "gridProperties": {
                            "rowCount": max(1, first_ws.max_row),
                            "columnCount": max(1, first_ws.max_column),
                        },
                    },
                    "fields": "title,index,gridProperties.rowCount,gridProperties.columnCount",
                }
            }
        )
    else:
        first_sheet_id = None

    for index, ws in enumerate(workbook.worksheets[1:], start=1):
        requests.append(
            {
                "addSheet": {
                    "properties": {
                        "title": ws.title,
                        "index": index,
                        "gridProperties": {
                            "rowCount": max(1, ws.max_row),
                            "columnCount": max(1, ws.max_column),
                        },
                    }
                }
            }
        )

    for stale in existing_sheets[1:]:
        requests.append({"deleteSheet": {"sheetId": int(stale["properties"]["sheetId"])}})

    if requests:
        sheets_service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"requests": requests},
        ).execute()
    return _get_spreadsheet_sheet_ids(sheets_service, spreadsheet_id)


def _create_spreadsheet(sheets_service, drive_service, title: str, workbook, folder_id: str | None) -> tuple[str, dict[str, int]]:
    if folder_id:
        body = {
            "name": title,
            "mimeType": "application/vnd.google-apps.spreadsheet",
            "parents": [folder_id],
        }
        response = drive_service.files().create(
            body=body,
            fields="id",
            supportsAllDrives=True,
        ).execute()
        spreadsheet_id = response["id"]
        sheet_ids = _configure_sheets_structure(sheets_service, spreadsheet_id, workbook)
        return spreadsheet_id, sheet_ids

    body = {
        "properties": {"title": title},
        "sheets": [
            {
                "properties": {
                    "title": ws.title,
                    "index": index,
                    "gridProperties": {
                        "rowCount": max(1, ws.max_row),
                        "columnCount": max(1, ws.max_column),
                    },
                }
            }
            for index, ws in enumerate(workbook.worksheets)
        ],
    }
    response = sheets_service.spreadsheets().create(body=body).execute()
    spreadsheet_id = response["spreadsheetId"]
    sheet_ids = {
        sheet["properties"]["title"]: int(sheet["properties"]["sheetId"])
        for sheet in response.get("sheets", [])
    }
    return spreadsheet_id, sheet_ids


def _move_to_drive_folder(drive_service, spreadsheet_id: str, folder_id: str) -> None:
    metadata = drive_service.files().get(fileId=spreadsheet_id, fields="parents").execute()
    parents = metadata.get("parents", [])
    drive_service.files().update(
        fileId=spreadsheet_id,
        addParents=folder_id,
        removeParents=",".join(parents) if parents else None,
        fields="id, parents",
    ).execute()


def export_workbook_to_gsheets(*, xlsx_path: Path, spreadsheet_title: str) -> tuple[str, str]:
    if not xlsx_path.exists() or not xlsx_path.is_file() or xlsx_path.suffix.lower() != ".xlsx":
        raise FileNotFoundError(f"Workbook not found or not an .xlsx file: {xlsx_path}")

    client_secret_path = _resolve_env_path("GOOGLE_OAUTH_CLIENT_SECRET_JSON")
    if client_secret_path is None:
        raise RuntimeError("Missing required environment variable: GOOGLE_OAUTH_CLIENT_SECRET_JSON")
    if not client_secret_path.exists() or not client_secret_path.is_file():
        raise FileNotFoundError(f"Google OAuth client secret file not found: {client_secret_path}")

    token_path = _resolve_env_path("GOOGLE_OAUTH_TOKEN_JSON", default="output/google_oauth_token.json")
    if token_path is None:
        raise RuntimeError("Could not resolve GOOGLE_OAUTH_TOKEN_JSON")

    credentials = _load_credentials(client_secret_path, token_path)
    sheets_service, drive_service = _build_services(credentials)

    workbook = load_workbook(xlsx_path, data_only=True)
    folder_id = (os.environ.get("GOOGLE_DRIVE_FOLDER_ID") or "").strip()
    spreadsheet_id, sheet_ids = _create_spreadsheet(
        sheets_service,
        drive_service,
        spreadsheet_title,
        workbook,
        folder_id or None,
    )
    if folder_id and not sheet_ids:
        _move_to_drive_folder(drive_service, spreadsheet_id, folder_id)

    for ws in workbook.worksheets:
        sheet_id = sheet_ids[ws.title]
        requests = _dimension_requests(ws, sheet_id)
        requests.append(_sheet_update_request(ws, sheet_id))
        requests.extend(_merge_requests(ws, sheet_id))
        for chunk in _chunked(requests, size=50):
            sheets_service.spreadsheets().batchUpdate(
                spreadsheetId=spreadsheet_id,
                body={"requests": chunk},
            ).execute()
        _safe_print(f"Uploaded sheet: {ws.title} ({ws.max_row} rows x {ws.max_column} cols)")

    spreadsheet_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}"
    return spreadsheet_id, spreadsheet_url


def main(argv: list[str] | None = None) -> int:
    load_dotenv(ROOT / ".env", override=False)
    parser = _build_parser()
    args = parser.parse_args(argv)

    xlsx_default = (os.environ.get("GOOGLE_SHEETS_SOURCE_XLSX") or "").strip() or str(DEFAULT_XLSX_PATH)
    xlsx_path = Path((args.xlsx_path or "").strip() or xlsx_default).expanduser().resolve()
    spreadsheet_title = (args.spreadsheet_title or "").strip() or _default_spreadsheet_title(xlsx_path)

    try:
        spreadsheet_id, spreadsheet_url = export_workbook_to_gsheets(
            xlsx_path=xlsx_path,
            spreadsheet_title=spreadsheet_title,
        )
    except HttpError as exc:
        raise RuntimeError(f"Google API request failed: {exc}") from exc

    _safe_print(f"Created spreadsheet ID: {spreadsheet_id}")
    _safe_print(f"Created spreadsheet URL: {spreadsheet_url}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
