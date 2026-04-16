from __future__ import annotations

import logging
import shutil
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


logger = logging.getLogger(__name__)
LOG_PREFIX = "[ Excel Layout: ]"


@dataclass(frozen=True)
class WorkbookContext:
    path: Path
    wb: Any
    data_sheet: Any        # single data sheet derived from the template
    headers: list[str]
    header_row_idx: int


FILE_INDEX_SHEET = "File Index"
PDF_SOURCE_HEADER = "pdf_source"


def _find_best_header_row(ws, *, max_scan_rows: int = 10) -> int:
    """
    Return the row index (1-based) that is most likely the header row.
    Heuristic: the row among the first max_scan_rows rows that has the
    most non-empty, non-purely-numeric cell values.
    """
    best_row = 1
    best_score = -1
    for row_idx in range(1, min(ws.max_row, max_scan_rows) + 1):
        cells = list(ws.iter_rows(min_row=row_idx, max_row=row_idx))[0]
        score = 0
        for cell in cells:
            v = cell.value
            if v is None:
                continue
            s = str(v).strip()
            if not s:
                continue
            try:
                float(s)
            except ValueError:
                score += 1
        if score > best_score:
            best_score = score
            best_row = row_idx
    return best_row


def find_data_sheet(wb) -> Any:
    """
    Return the first worksheet in the workbook that is not the File Index
    and has at least one non-empty header cell.

    Raises RuntimeError if no suitable sheet is found.
    """
    for ws in wb.worksheets:
        if ws.title == FILE_INDEX_SHEET:
            continue
        if ws.max_row >= 1 and ws.max_column >= 1:
            header_row_idx = _find_best_header_row(ws)
            row_cells = list(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx))[0]
            headers = [str(c.value).strip() for c in row_cells if c.value is not None and str(c.value).strip()]
            if headers:
                return ws
    raise RuntimeError(
        "Workbook has no suitable data sheet. "
        "Ensure the template contains at least one sheet (other than 'File Index') with header rows."
    )


def load_workbook_context(xlsx_path: Path) -> WorkbookContext:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = find_data_sheet(wb)
    header_row_idx = _find_best_header_row(ws)
    row_cells = list(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx))[0]
    headers = [str(c.value).strip() for c in row_cells if c.value is not None and str(c.value).strip()]
    if not headers:
        raise RuntimeError(f"Sheet '{ws.title}' has no detectable headers.")
    return WorkbookContext(path=xlsx_path, wb=wb, data_sheet=ws, headers=headers, header_row_idx=header_row_idx)


def write_timestamped_copy(xlsx_path: Path) -> Path:
    ts = datetime.now().strftime("%m%d_%H%M")
    versioned_path = xlsx_path.with_name(f"{xlsx_path.stem}.{ts}{xlsx_path.suffix}")
    shutil.copy2(xlsx_path, versioned_path)
    return versioned_path


def ensure_file_index_sheet(wb) -> Any:
    if FILE_INDEX_SHEET in wb.sheetnames:
        return wb[FILE_INDEX_SHEET]
    ws = wb.create_sheet(FILE_INDEX_SHEET)
    ws.append(["Ref #", "Relative PDF Path", "SHA256", "Chars Extracted", "Model", "Processed At (ISO)", PDF_SOURCE_HEADER])
    return ws


def _find_row_by_ref(ws, ref_number: int, ref_col_index: int, *, start_row: int = 2) -> int | None:
    for row_idx in range(start_row, ws.max_row + 1):
        v = ws.cell(row=row_idx, column=ref_col_index).value
        if v is None:
            continue
        try:
            if int(v) == int(ref_number):
                return row_idx
        except Exception:
            continue
    return None


def _find_file_index_row_by_sha(file_index_ws, sha256: str) -> int | None:
    sha_col = 3  # File Index columns are fixed: Ref #, Path, SHA256, ...
    for row_idx in range(2, file_index_ws.max_row + 1):
        v = file_index_ws.cell(row=row_idx, column=sha_col).value
        if v and str(v).strip() == sha256:
            return row_idx
    return None


def _clear_row(ws, row_idx: int, *, max_col: int) -> None:
    for col_idx in range(1, max_col + 1):
        ws.cell(row=row_idx, column=col_idx).value = None


MAX_OUTPUT_COLUMN_WIDTH = 85.0


def apply_output_sheet_layout(ws, *, min_row: int = 1, max_width: float = MAX_OUTPUT_COLUMN_WIDTH, padding: float = 2.0) -> None:
    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        existing_width = ws.column_dimensions[column_letter].width
        target_width = float(existing_width) if existing_width else 0.0

        for row_idx in range(min_row, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            text = str(value)
            if not text:
                continue
            longest_line = max(len(line) for line in text.splitlines()) if text else 0
            target_width = max(target_width, min(float(max_width), float(longest_line) + float(padding)))

        if target_width > 0:
            ws.column_dimensions[column_letter].width = target_width

    for row_idx in range(min_row, ws.max_row + 1):
        row_dim = ws.row_dimensions[row_idx]
        row_dim.height = None
        if hasattr(row_dim, "ht"):
            try:
                row_dim.ht = None
            except Exception:
                pass
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, str):
                cell.number_format = "@"
            base_alignment = cell.alignment or Alignment()
            cell.alignment = copy_alignment_with_wrap(base_alignment)


def autofit_workbook_with_excel(xlsx_path: Path, *, max_width: float = MAX_OUTPUT_COLUMN_WIDTH) -> None:
    try:
        import pythoncom
        import win32com.client
    except Exception:
        logger.info("%s Excel COM autofit unavailable; skipped Excel-native autofit for %s", LOG_PREFIX, xlsx_path)
        return

    pythoncom.CoInitialize()
    excel = None
    workbook = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(str(xlsx_path.resolve()))

        for worksheet in workbook.Worksheets:
            used_range = worksheet.UsedRange
            used_range.WrapText = True
            used_range.Columns.AutoFit()

            for column in used_range.Columns:
                if float(column.ColumnWidth) > float(max_width):
                    column.ColumnWidth = float(max_width)

            used_range.Rows.AutoFit()

        workbook.Save()
        logger.info("%s Applied Excel-native autofit to %s", LOG_PREFIX, xlsx_path)
    except Exception as e:
        logger.warning(
            "%s Excel COM autofit failed for %s: %s: %s",
            LOG_PREFIX,
            xlsx_path,
            type(e).__name__,
            e,
        )
    finally:
        if workbook is not None:
            workbook.Close(SaveChanges=False)
        if excel is not None:
            excel.Quit()
        pythoncom.CoUninitialize()


def copy_alignment_with_wrap(alignment: Alignment) -> Alignment:
    return Alignment(
        horizontal=alignment.horizontal,
        vertical=alignment.vertical or "top",
        textRotation=alignment.textRotation,
        wrapText=True,
        shrinkToFit=alignment.shrinkToFit,
        indent=alignment.indent,
        relativeIndent=alignment.relativeIndent,
        justifyLastLine=alignment.justifyLastLine,
        readingOrder=alignment.readingOrder,
    )


def write_paper_row(
    ctx: WorkbookContext,
    *,
    row_data: dict[str, str],
    file_index: dict[str, Any],
    overwrite_existing: bool,
):
    headers = ctx.headers
    ws = ctx.data_sheet
    file_ws = ensure_file_index_sheet(ctx.wb)

    # Use column 1 as the ref/ID column regardless of its name
    ref_col_index = 1
    sha256 = str(file_index["sha256"])

    try:
        ref_number = int(row_data.get(headers[0]) or 0) if headers else 0
    except (TypeError, ValueError):
        ref_number = 0

    target_row_idx: int | None = None
    existing_ref_int: int | None = None

    if overwrite_existing:
        existing_file_row = _find_file_index_row_by_sha(file_ws, sha256)
        if existing_file_row is not None:
            existing_ref = file_ws.cell(row=existing_file_row, column=1).value
            try:
                existing_ref_int = int(existing_ref)
            except Exception:
                existing_ref_int = ref_number
        if existing_ref_int is None and ref_number:
            existing_ref_int = ref_number

        if existing_ref_int:
            target_row_idx = _find_row_by_ref(ws, existing_ref_int, ref_col_index, start_row=ctx.header_row_idx + 1)
        if target_row_idx is None and ref_number:
            target_row_idx = _find_row_by_ref(ws, ref_number, ref_col_index, start_row=ctx.header_row_idx + 1)

    if target_row_idx is None:
        target_row_idx = ws.max_row + 1

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=target_row_idx, column=col_idx).value = row_data.get(header)

    # Update File Index (append or overwrite by SHA)
    processed_at = datetime.now(timezone.utc).isoformat()
    existing_file_row = _find_file_index_row_by_sha(file_ws, sha256)

    if existing_file_row is None:
        file_ws.append(
            [
                ref_number,
                file_index["relative_path"],
                sha256,
                file_index["chars_extracted"],
                file_index["model"],
                processed_at,
                file_index.get("pdf_source", ""),
            ]
        )
    else:
        file_ws.cell(row=existing_file_row, column=1).value = ref_number
        file_ws.cell(row=existing_file_row, column=2).value = file_index["relative_path"]
        file_ws.cell(row=existing_file_row, column=4).value = file_index["chars_extracted"]
        file_ws.cell(row=existing_file_row, column=5).value = file_index["model"]
        file_ws.cell(row=existing_file_row, column=6).value = processed_at
        file_ws.cell(row=existing_file_row, column=7).value = file_index.get("pdf_source", "")


def save_workbook(ctx: WorkbookContext):
    apply_output_sheet_layout(ctx.data_sheet, min_row=1)
    file_ws = ensure_file_index_sheet(ctx.wb)
    apply_output_sheet_layout(file_ws, min_row=1)
    ctx.wb.save(ctx.path)
    autofit_workbook_with_excel(ctx.path)
    write_timestamped_copy(ctx.path)
