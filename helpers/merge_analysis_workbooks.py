from __future__ import annotations

import argparse
import os
import sys
from copy import copy
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

try:
    from helpers.env import load_dotenv
    from helpers.excel_writer import (
        FILE_INDEX_SHEET,
        apply_output_sheet_layout,
        autofit_workbook_with_excel,
        write_timestamped_copy,
    )
    from helpers.logging_utils import get_logger
except ModuleNotFoundError:
    from helpers.env import load_dotenv
    from helpers.excel_writer import (
        FILE_INDEX_SHEET,
        apply_output_sheet_layout,
        autofit_workbook_with_excel,
        write_timestamped_copy,
    )
    from helpers.logging_utils import get_logger

DEFAULT_PAPER_XLSX = ROOT / "output" / "paper_wise_analysis" / "paper_analysis.xlsx"
DEFAULT_INGREDIENT_XLSX = ROOT / "output" / "ingredient_wise_analysis" / "matrix_weight_management.populated.xlsx"
DEFAULT_OUTPUT_XLSX = ROOT / "output" / "ingredient_analysis_output.xlsx"
DEFAULT_PAPER_DIR = DEFAULT_PAPER_XLSX.parent
DEFAULT_INGREDIENT_DIR = DEFAULT_INGREDIENT_XLSX.parent
logger = get_logger(__name__, prefix="[ Workbook Merge: ]")


def _iter_candidate_workbooks(directory: Path) -> list[Path]:
    if not directory.exists() or not directory.is_dir():
        return []
    return sorted(
        (
            path for path in directory.glob("*.xlsx")
            if path.is_file() and not path.name.startswith("~$")
        ),
        key=lambda path: (path.stat().st_mtime, path.name.lower()),
        reverse=True,
    )


def _resolve_input_path(
    *,
    explicit_value: str,
    env_var: str,
    default_path: Path,
    search_dir: Path,
    label: str,
) -> Path:
    raw_value = explicit_value.strip()
    if raw_value:
        return Path(raw_value).expanduser().resolve()

    env_value = (os.environ.get(env_var) or "").strip()
    if env_value:
        env_path = Path(env_value).expanduser().resolve()
        if env_path.exists():
            return env_path
        logger.warning(
            "Configured %s path does not exist: %s. Falling back to workbook discovery in %s.",
            env_var,
            env_path,
            search_dir,
        )

    if default_path.exists():
        return default_path.resolve()

    candidates = _iter_candidate_workbooks(search_dir)
    if candidates:
        chosen = candidates[0].resolve()
        logger.info("Resolved %s workbook from latest file in %s: %s", label, search_dir, chosen)
        return chosen

    return default_path.resolve()


def _available_workbooks_message(directory: Path) -> str:
    candidates = _iter_candidate_workbooks(directory)
    if not candidates:
        return f"No .xlsx files found under {directory}"
    return "Available .xlsx files: " + ", ".join(str(path.name) for path in candidates[:5])


def copy_sheet(source: Worksheet, target: Worksheet) -> None:
    target.sheet_format.defaultColWidth = source.sheet_format.defaultColWidth
    target.sheet_format.defaultRowHeight = source.sheet_format.defaultRowHeight
    target.sheet_format.zeroHeight = source.sheet_format.zeroHeight
    target.sheet_properties = copy(source.sheet_properties)
    target.page_margins = copy(source.page_margins)
    target.page_setup = copy(source.page_setup)
    target.print_options = copy(source.print_options)
    target.freeze_panes = source.freeze_panes
    if source.auto_filter and source.auto_filter.ref:
        target.auto_filter.ref = source.auto_filter.ref
    try:
        target.print_title_cols = source.print_title_cols
    except Exception:
        pass
    try:
        target.print_title_rows = source.print_title_rows
    except Exception:
        pass
    try:
        target.sheet_view.selection = copy(source.sheet_view.selection)
    except Exception:
        pass
    target.sheet_view.showGridLines = source.sheet_view.showGridLines
    target.sheet_view.showRowColHeaders = source.sheet_view.showRowColHeaders
    target.sheet_view.zoomScale = source.sheet_view.zoomScale
    target.sheet_view.zoomScaleNormal = source.sheet_view.zoomScaleNormal
    target.sheet_view.tabSelected = source.sheet_view.tabSelected
    target.sheet_view.topLeftCell = source.sheet_view.topLeftCell

    for col_key, dim in source.column_dimensions.items():
        target_dim = target.column_dimensions[col_key]
        for attr in (
            "width",
            "bestFit",
            "hidden",
            "outlineLevel",
            "outline_level",
            "collapsed",
            "style",
            "min",
            "max",
        ):
            if hasattr(dim, attr):
                try:
                    setattr(target_dim, attr, getattr(dim, attr))
                except Exception:
                    pass

    for row_idx, dim in source.row_dimensions.items():
        target_dim = target.row_dimensions[row_idx]
        for attr in (
            "height",
            "hidden",
            "outlineLevel",
            "outline_level",
            "collapsed",
            "style",
            "ht",
        ):
            if hasattr(dim, attr):
                try:
                    setattr(target_dim, attr, getattr(dim, attr))
                except Exception:
                    pass

    for merged_range in source.merged_cells.ranges:
        target.merge_cells(str(merged_range))

    for row in source.iter_rows():
        for cell in row:
            new_cell = target.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell._style = copy(cell._style)
            if cell.number_format:
                new_cell.number_format = cell.number_format
            if cell.font:
                new_cell.font = copy(cell.font)
            if cell.fill:
                new_cell.fill = copy(cell.fill)
            if cell.border:
                new_cell.border = copy(cell.border)
            if cell.alignment:
                new_cell.alignment = copy(cell.alignment)
            if cell.protection:
                new_cell.protection = copy(cell.protection)
            if cell.hyperlink:
                new_cell._hyperlink = copy(cell.hyperlink)
            if cell.comment:
                new_cell.comment = copy(cell.comment)


def validate_input(path: Path, label: str) -> None:
    if not path.exists() or not path.is_file() or path.suffix.lower() != ".xlsx":
        search_dir = path.parent
        raise FileNotFoundError(
            f"{label} workbook not found or not an .xlsx file: {path}. "
            f"{_available_workbooks_message(search_dir)}"
        )


def merge_workbooks(paper_xlsx: Path, ingredient_xlsx: Path, output_xlsx: Path) -> None:
    validate_input(paper_xlsx, "Paper")
    validate_input(ingredient_xlsx, "Ingredient")

    paper_wb = load_workbook(paper_xlsx)
    ingredient_wb = load_workbook(ingredient_xlsx)

    try:
        if not ingredient_wb.worksheets:
            raise RuntimeError(f"Ingredient workbook has no worksheets: {ingredient_xlsx}")

        output_xlsx.parent.mkdir(parents=True, exist_ok=True)

        merged_wb = Workbook()
        default_sheet = merged_wb.active
        merged_wb.remove(default_sheet)

        # Copy all sheets from the ingredient workbook first.
        for ws_src in ingredient_wb.worksheets:
            target_title = ws_src.title if ws_src.title != "Sheet" else "Ingredient Matrix"
            target = merged_wb.create_sheet(title=target_title)
            copy_sheet(ws_src, target)
            apply_output_sheet_layout(target, min_row=1)

        # Copy all sheets from the paper workbook (data sheet(s) first, then File Index).
        data_sheets = [ws for ws in paper_wb.worksheets if ws.title != FILE_INDEX_SHEET]
        index_sheets = [ws for ws in paper_wb.worksheets if ws.title == FILE_INDEX_SHEET]

        for ws_src in data_sheets + index_sheets:
            target = merged_wb.create_sheet(title=ws_src.title)
            copy_sheet(ws_src, target)
            apply_output_sheet_layout(target, min_row=1)

        tmp_output = output_xlsx.with_suffix(output_xlsx.suffix + ".tmp")
        if tmp_output.exists():
            tmp_output.unlink()
        merged_wb.save(tmp_output)
        tmp_output.replace(output_xlsx)
        autofit_workbook_with_excel(output_xlsx)
        write_timestamped_copy(output_xlsx)
    finally:
        paper_wb.close()
        ingredient_wb.close()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(prog="merge_analysis_workbooks")
    parser.add_argument(
        "--paper-xlsx",
        default="",
        help="Path to the paper analysis workbook. If omitted, uses PAPER_EXTRACT_OUTPUT_XLSX, "
        "output/paper_wise_analysis/paper_analysis.xlsx, or the newest .xlsx in that folder.",
    )
    parser.add_argument(
        "--ingredient-xlsx",
        default="",
        help="Path to the ingredient matrix workbook. If omitted, uses INGREDIENT_MATRIX_OUTPUT_XLSX, "
        "output/ingredient_wise_analysis/matrix_weight_management.populated.xlsx, or the newest .xlsx in that folder.",
    )
    parser.add_argument(
        "--output-xlsx",
        default="",
        help="Path for the merged workbook.",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    load_dotenv(ROOT / ".env", override=False)
    parser = build_parser()
    args = parser.parse_args(argv)

    paper_xlsx = _resolve_input_path(
        explicit_value=args.paper_xlsx,
        env_var="PAPER_EXTRACT_OUTPUT_XLSX",
        default_path=DEFAULT_PAPER_XLSX,
        search_dir=DEFAULT_PAPER_DIR,
        label="paper",
    )
    ingredient_xlsx = _resolve_input_path(
        explicit_value=args.ingredient_xlsx,
        env_var="INGREDIENT_MATRIX_OUTPUT_XLSX",
        default_path=DEFAULT_INGREDIENT_XLSX,
        search_dir=DEFAULT_INGREDIENT_DIR,
        label="ingredient",
    )
    output_value = args.output_xlsx.strip() or (os.environ.get("MERGED_ANALYSIS_OUTPUT_XLSX") or "").strip()
    output_xlsx = Path(output_value or DEFAULT_OUTPUT_XLSX).expanduser().resolve()

    merge_workbooks(paper_xlsx, ingredient_xlsx, output_xlsx)
    logger.info("Created merged workbook: %s", output_xlsx)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
