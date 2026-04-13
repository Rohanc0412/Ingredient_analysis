from __future__ import annotations

import json
import os
import shutil
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

from helpers.excel_writer import apply_output_sheet_layout, autofit_workbook_with_excel, write_timestamped_copy
from helpers.env import load_dotenv
from helpers.file_discovery import sorted_glob_files
from helpers.logging_utils import get_logger


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_INPUT_DIR = ROOT / "input" / "ingredient_analysis"
DEFAULT_TEMPLATE_PATH = ROOT / "input" / "templates" / "matrix_weight_management.template.xlsx"
DEFAULT_OUTPUT_PATH = ROOT / "output" / "ingredient_wise_analysis" / "matrix_weight_management.populated.xlsx"
logger = get_logger(__name__, prefix="[ Ingredient Matrix: ]")


def normalize_key(value: str) -> str:
    text = unicodedata.normalize("NFKC", value)
    replacements = {
        "\u2010": "-",
        "\u2011": "-",
        "\u2012": "-",
        "\u2013": "-",
        "\u2014": "-",
        "\u2212": "-",
        "\u00a0": " ",
    }
    for src, dst in replacements.items():
        text = text.replace(src, dst)
    return " ".join(text.split()).strip().casefold()


def _stringify_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, dict):
        return extract_answer(value)
    if isinstance(value, list):
        parts = [_stringify_value(item).strip() for item in value]
        return "\n".join(part for part in parts if part)
    return str(value)


def extract_answer(value) -> str:
    if value is None:
        return ""
    if isinstance(value, dict):
        ignored_keys = {"sources", "source", "citations", "citation", "evidence", "metadata"}
        preferred_order = ("answer", "category", "summary", "items", "value", "unit", "context", "study_count")
        ignored_normalized = {normalize_key(k) for k in ignored_keys}
        preferred_normalized = {normalize_key(k) for k in preferred_order}
        parts: list[str] = []
        seen: set[str] = set()

        def add_part(text: str) -> None:
            cleaned = " ".join(str(text).split()).strip() if "\n" not in str(text) else str(text).strip()
            if not cleaned or cleaned in seen:
                return
            seen.add(cleaned)
            parts.append(cleaned)

        scalar_bundle: dict[str, str] = {}
        for key in preferred_order:
            item = value.get(key)
            if isinstance(item, (dict, list)):
                continue
            text = _stringify_value(item).strip()
            if text:
                scalar_bundle[key] = text

        if scalar_bundle.get("value"):
            dose = " ".join(part for part in (scalar_bundle.get("value", ""), scalar_bundle.get("unit", "")) if part).strip()
            context = scalar_bundle.get("context", "")
            if dose and context:
                add_part(f"{dose}\n{context}")
            elif dose:
                add_part(dose)
            elif context:
                add_part(context)
            scalar_bundle.pop("value", None)
            scalar_bundle.pop("unit", None)
            scalar_bundle.pop("context", None)

        for key in preferred_order:
            text = scalar_bundle.get(key)
            if text:
                add_part(text)

        for key in preferred_order:
            item = value.get(key)
            if not isinstance(item, (dict, list)):
                continue
            text = _stringify_value(item).strip()
            if text:
                add_part(text)

        for key, item in value.items():
            normalized_key = normalize_key(str(key))
            if normalized_key in ignored_normalized or normalized_key in preferred_normalized:
                continue
            text = _stringify_value(item).strip()
            if text:
                add_part(text)

        return "\n".join(parts)
    if isinstance(value, list):
        return "\n".join(part for part in (_stringify_value(item).strip() for item in value) if part)
    return str(value)


def load_records(input_dir: Path) -> list[dict[str, str]]:
    records: list[dict[str, str]] = []
    for path in sorted_glob_files(input_dir, "*.json"):
        payload = json.loads(path.read_text(encoding="utf-8"))
        fields = payload.get("fields", {})
        record = {
            normalize_key(key): extract_answer(value)
            for key, value in fields.items()
        }
        ingredient = payload.get("ingredient")
        if ingredient and not record.get(normalize_key("Ingredient")):
            record[normalize_key("Ingredient")] = str(ingredient)
        records.append(record)
    return records


def populate_workbook(template_path: Path, output_path: Path, records: list[dict[str, str]]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, output_path)

    wb = load_workbook(output_path)
    ws = wb[wb.sheetnames[0]]

    headers = [ws.cell(1, col).value or "" for col in range(1, ws.max_column + 1)]
    normalized_headers = [normalize_key(str(header)) for header in headers]

    start_row = 2
    for row_offset, record in enumerate(records):
        row_idx = start_row + row_offset
        for col_idx, header_key in enumerate(normalized_headers, start=1):
            ws.cell(row_idx, col_idx).value = record.get(header_key, "")

    apply_output_sheet_layout(ws, min_row=1)
    wb.save(output_path)
    autofit_workbook_with_excel(output_path)
    write_timestamped_copy(output_path)


def main() -> int:
    load_dotenv(ROOT / ".env", override=False)

    input_dir = Path(os.environ.get("INGREDIENT_MATRIX_INPUT_DIR") or DEFAULT_INPUT_DIR).expanduser().resolve()
    template_path = Path(os.environ.get("INGREDIENT_MATRIX_TEMPLATE_XLSX") or DEFAULT_TEMPLATE_PATH).expanduser().resolve()
    output_path = Path(os.environ.get("INGREDIENT_MATRIX_OUTPUT_XLSX") or DEFAULT_OUTPUT_PATH).expanduser().resolve()

    if not template_path.exists():
        raise FileNotFoundError(f"Template workbook not found: {template_path}")
    if not input_dir.exists():
        raise FileNotFoundError(f"Input directory not found: {input_dir}")

    records = load_records(input_dir)
    if not records:
        raise RuntimeError(f"No JSON files found in {input_dir}")

    populate_workbook(template_path, output_path, records)
    logger.info("Created workbook: %s", output_path)
    logger.info("Rows populated: %s", len(records))
    return 0
